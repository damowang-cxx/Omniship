from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy import select

from app.db.models import BillingEntry, WaybillUpload
from app.services.supplier_defaults import QLS_SUPPLIER_ID, QLS_VERSION_ID
from tests.auth_helpers import create_test_user, login


def workbook_bytes(parcel_values: list[str], *, invalid_quantity: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([f"Column {index}" for index in range(1, 24)])
    for index, parcel in enumerate(parcel_values):
        row = [""] * 23
        row[8] = parcel
        row[18] = "DE"
        row[20] = 21 if invalid_quantity and index == 0 else 1
        row[21] = 0.5
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def rpl_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Instructions"
    sheet = workbook.create_sheet("Shipments")
    sheet.append(["RPL export"])
    sheet.append(["Generated for integration test"])
    sheet.append(["Tracking Reference", "Qty"])
    sheet.append(["AB-1", 1])
    sheet.append([" ab-1 ", "not-an-integer"])
    sheet.append(["XY-2", 5])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def upload_files(content: bytes):
    return [
        ("airWaybillDocuments", ("awb.pdf", b"%PDF-1.4\ncontent", "application/pdf")),
        (
            "preAlertFile",
            (
                "pre-alert.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
    ]


def upload_data(*, airport: str = "AMS", number: str = "784-84063276"):
    return {
        "shipmentType": "Air",
        "airWaybillNumber": number,
        "grossWeightKg": "12.5",
        "pieces": "8",
        "arrivalFlightNumber": "EK0147",
        "airportOfDeparture": "HKG",
        "airportOfArrival": airport,
        "supplierId": str(QLS_SUPPLIER_ID),
    }


def rpl_config():
    return {
        "workbook": {
            "sheetMode": "named",
            "sheetName": "Shipments",
            "headerRow": 3,
            "dataStartRow": 4,
        },
        "fields": [
            {
                "key": "tracking_reference",
                "name": "Tracking Reference",
                "semanticField": None,
                "locatorMode": "header",
                "locatorValue": "Tracking Reference",
                "valueType": "text",
                "blankPolicy": "skip_row",
                "caseInsensitive": True,
                "allowUnknownCountry": True,
                "countryAliases": {},
                "constraints": {"allowedValues": [], "unique": False},
            },
            {
                "key": "quantity",
                "name": "Quantity",
                "semanticField": None,
                "locatorMode": "header",
                "locatorValue": "Qty",
                "valueType": "integer",
                "blankPolicy": "allow",
                "caseInsensitive": False,
                "allowUnknownCountry": True,
                "countryAliases": {},
                "constraints": {
                    "minValue": "0",
                    "maxValue": "20",
                    "allowedValues": [],
                    "unique": False,
                },
            },
        ],
        "rowKeyFieldKey": "tracking_reference",
        "billingDistinctFieldKey": "tracking_reference",
    }


def test_warning_upload_deducts_distinct_units_atomically(client, db_session):
    user = create_test_user(
        db_session,
        email="user@example.com",
        username="User",
        balance="20.00",
    )
    assert login(client, email=user.email).status_code == 200

    response = client.post(
        "/api/v1/waybill-uploads/file",
        data=upload_data(),
        files=upload_files(
            workbook_bytes(["PARCEL-1", "PARCEL-1", "PARCEL-2", "PARCEL-3", "PARCEL-4", "PARCEL-5"], invalid_quantity=True)
        ),
    )

    assert response.status_code == 201
    body = response.json()
    assert body["billableUnitCount"] == 5
    assert body["deductedTax"] == "15.00"
    assert body["balanceAfter"] == "5.00"
    assert body["validationIssueCount"] == 1
    assert "less than or equal to 20" in body["validationIssues"][0]["message"]

    entry = db_session.execute(select(BillingEntry)).scalar_one()
    assert entry.entry_type == "deduction"
    assert entry.amount == Decimal("15.00")
    assert entry.billable_unit_count == 5
    assert entry.supplier_name == "QLS"


def test_insufficient_balance_saves_nothing(client, db_session):
    user = create_test_user(
        db_session,
        email="user@example.com",
        username="User",
        balance="10.00",
    )
    assert login(client, email=user.email).status_code == 200

    response = client.post(
        "/api/v1/waybill-uploads/file",
        data=upload_data(),
        files=upload_files(workbook_bytes(["P1", "P2", "P3", "P4", "P5"])),
    )

    assert response.status_code == 402
    assert response.json()["detail"] == {
        "code": "insufficient_balance",
        "required": "15.00",
        "balance": "10.00",
        "shortfall": "5.00",
    }
    assert db_session.execute(select(WaybillUpload)).scalars().all() == []
    assert db_session.execute(select(BillingEntry)).scalars().all() == []
    db_session.refresh(user)
    assert user.balance == Decimal("10.00")


def test_non_taxable_airport_uploads_without_deduction(client, db_session):
    user = create_test_user(
        db_session,
        email="user@example.com",
        username="User",
        balance="10.00",
    )
    assert login(client, email=user.email).status_code == 200

    response = client.post(
        "/api/v1/waybill-uploads/file",
        data=upload_data(airport="LHR"),
        files=upload_files(workbook_bytes(["P1", "P2", "P3"])),
    )

    assert response.status_code == 201
    assert response.json()["deductedTax"] == "0.00"
    assert response.json()["balanceAfter"] == "10.00"
    assert db_session.execute(select(BillingEntry)).scalars().all() == []


def test_zero_distinct_values_is_structural_error(client, db_session):
    user = create_test_user(db_session, email="user@example.com", username="User")
    assert login(client, email=user.email).status_code == 200

    response = client.post(
        "/api/v1/waybill-uploads/file",
        data=upload_data(),
        files=upload_files(workbook_bytes([])),
    )

    assert response.status_code == 400
    assert "contains no valid values" in response.text


def test_header_based_supplier_with_named_sheet_needs_no_code_change(client, db_session):
    admin = create_test_user(
        db_session,
        email="admin@example.com",
        username="Admin",
        role="admin",
    )
    user = create_test_user(
        db_session,
        email="user@example.com",
        username="User",
        balance="20.00",
    )
    assert login(client, email=admin.email).status_code == 200
    config = rpl_config()
    supplier_response = client.post(
        "/api/v1/suppliers",
        json={"name": "RPL", "config": config},
    )
    assert supplier_response.status_code == 201
    supplier_id = supplier_response.json()["id"]

    assert login(client, email=user.email).status_code == 200
    content = rpl_workbook_bytes()
    estimate = client.post(
        "/api/v1/billing/estimate",
        data={"supplierId": supplier_id, "airportOfArrival": "AMS"},
        files={
            "preAlertFile": (
                "rpl.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert estimate.status_code == 200
    assert estimate.json()["billableUnitCount"] == 2
    assert estimate.json()["estimatedTax"] == "6.00"
    assert estimate.json()["warningCount"] == 1

    payload = upload_data()
    payload["supplierId"] = supplier_id
    upload = client.post(
        "/api/v1/waybill-uploads/file",
        data=payload,
        files=upload_files(content),
    )
    assert upload.status_code == 201
    assert upload.json()["supplierName"] == "RPL"
    assert upload.json()["billableUnitCount"] == 2
    assert upload.json()["validationIssueCount"] == 1


def test_retroactive_billing_detects_supplier_and_allows_negative_balance(client, db_session):
    admin = create_test_user(
        db_session,
        email="admin@example.com",
        username="Admin",
        role="admin",
    )
    user = create_test_user(
        db_session,
        email="user@example.com",
        username="User",
        balance="0.00",
    )
    assert login(client, email=admin.email).status_code == 200
    supplier_response = client.post(
        "/api/v1/suppliers",
        json={"name": "RPL", "config": rpl_config()},
    )
    assert supplier_response.status_code == 201
    rpl_supplier = supplier_response.json()

    assert login(client, email=user.email).status_code == 200
    content = rpl_workbook_bytes()
    uploaded = client.post(
        "/api/v1/waybill-uploads/file",
        data={
            **upload_data(airport="LHR", number="020-12345678"),
            "supplierId": rpl_supplier["id"],
        },
        files=upload_files(content),
    )
    assert uploaded.status_code == 201
    upload = db_session.get(WaybillUpload, UUID(uploaded.json()["uploadId"]))
    upload.airport_of_arrival = "AMS"
    upload.supplier_id = QLS_SUPPLIER_ID
    upload.supplier_version_id = QLS_VERSION_ID
    db_session.commit()

    assert login(client, email=admin.email).status_code == 200
    assert client.patch(
        f"/api/v1/waybill-uploads/{upload.id}/status",
        json={"status": "approved"},
    ).status_code == 200
    response = client.post(
        "/api/v1/billing/retroactive",
        json={"waybillNumbers": ["020-12345678", "MISSING-1"]},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["succeededCount"] == 1
    assert body["failedCount"] == 1
    assert body["succeeded"][0]["supplierName"] == "RPL"
    assert body["succeeded"][0]["billableUnitCount"] == 2
    assert body["succeeded"][0]["amount"] == "6.00"
    assert body["succeeded"][0]["balanceAfter"] == "-6.00"
    assert body["failed"][0]["waybillNumber"] == "MISSING-1"

    account = client.get(f"/api/v1/billing/users/{user.id}")
    assert account.status_code == 200
    assert account.json()["user"]["balance"] == "-6.00"
    assert account.json()["deductions"][0]["billingSource"] == "retroactive"

    waybills = client.get("/api/v1/waybills")
    assert waybills.status_code == 200
    assert waybills.json()["items"][0]["customsCartons"] == 2
    assert waybills.json()["items"][0]["customsAmount"] == "6.00"

    retry = client.post(
        "/api/v1/billing/retroactive",
        json={"waybillNumbers": ["020-12345678"]},
    )
    assert retry.status_code == 200
    assert retry.json()["succeededCount"] == 0
    assert "already" in retry.json()["failed"][0]["reason"].lower()


def test_retroactive_billing_respects_july_first_effective_date(client, db_session):
    admin = create_test_user(
        db_session,
        email="admin@example.com",
        username="Admin",
        role="admin",
    )
    user = create_test_user(
        db_session,
        email="user@example.com",
        username="User",
        balance="0.00",
    )
    assert login(client, email=user.email).status_code == 200
    uploaded = client.post(
        "/api/v1/waybill-uploads/file",
        data=upload_data(airport="LHR", number="784-00000001"),
        files=upload_files(workbook_bytes(["P1", "P2"])),
    )
    assert uploaded.status_code == 201
    upload = db_session.get(WaybillUpload, UUID(uploaded.json()["uploadId"]))
    upload.airport_of_arrival = "AMS"
    upload.status = "approved"
    upload.created_at = datetime(2026, 6, 30, 23, 59, tzinfo=timezone.utc)
    db_session.commit()

    assert login(client, email=admin.email).status_code == 200
    response = client.post(
        "/api/v1/billing/retroactive",
        json={"waybillNumbers": ["784-00000001"]},
    )

    assert response.status_code == 200
    assert response.json()["succeededCount"] == 0
    assert "2026-07-01" in response.json()["failed"][0]["reason"]
    db_session.refresh(user)
    assert user.balance == Decimal("0.00")
    assert db_session.execute(select(BillingEntry)).scalars().all() == []
