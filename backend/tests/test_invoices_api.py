from datetime import date
from io import BytesIO
from zipfile import ZipFile

from PIL import Image
from openpyxl import load_workbook
from sqlalchemy import select

from app.db.models import BillingEntry, Invoice, InvoiceSettings
from app.services.invoice_export_service import _compact_stamp_anchor
from tests.auth_helpers import create_test_user, login


def stamp_png() -> bytes:
    image = Image.new("RGBA", (24, 24), (190, 20, 40, 255))
    output = BytesIO()
    image.save(output, "PNG")
    return output.getvalue()


def ready_settings(db_session, tmp_path, admin):
    stamp_path = tmp_path / "stamp.png"
    stamp_path.write_bytes(stamp_png())
    settings = InvoiceSettings(
        id=1,
        issuer_company_name="EPIX LOGISTICS CO., LIMITED",
        issuer_address_info="Issuer address",
        beneficiary_name="Epix Logistics Co., Limited",
        bank_account="7949929686",
        bank_name_and_code="DBS (016)",
        branch_code="478",
        swift_bic="DHBKHKHH",
        bank_address="Hong Kong",
        stamp_storage_path=str(stamp_path),
        stamp_original_filename="stamp.png",
        updated_by_user_id=admin.id,
    )
    db_session.add(settings)


def deduction(db_session, user, admin, index: int) -> BillingEntry:
    entry = BillingEntry(
        user_id=user.id,
        entry_type="deduction",
        amount="3.00",
        currency="EUR",
        balance_after="100.00",
        waybill_number=f"217-000{index:05d}",
        billable_unit_count=1,
        unit_rate="3.00",
        billing_source="upload",
        created_by_user_id=admin.id,
    )
    db_session.add(entry)
    return entry


def test_customer_can_create_invoice_export_and_admin_can_void(client, db_session, tmp_path):
    admin = create_test_user(db_session, email="admin@example.com", username="Admin", role="admin")
    user = create_test_user(db_session, email="customer@example.com", username="Customer")
    user.payer_company_name = "Customer Limited"
    user.payer_address_info = "Customer address"
    ready_settings(db_session, tmp_path, admin)
    first = deduction(db_session, user, admin, 1)
    second = deduction(db_session, user, admin, 2)
    db_session.commit()

    assert login(client, email=user.email).status_code == 200
    eligible = client.get("/api/v1/billing/me/invoices/eligible")
    assert eligible.status_code == 200
    assert {item["id"] for item in eligible.json()} == {str(first.id), str(second.id)}

    created = client.post(
        "/api/v1/billing/me/invoices",
        json={"deductionIds": [str(first.id), str(second.id)], "issuedDate": "2026-08-17"},
    )
    assert created.status_code == 201
    invoice = created.json()["invoices"][0]
    assert invoice["invoiceNumber"] == "INV26001"
    assert invoice["dueDate"] == "2026-08-21"
    assert invoice["totalAmount"] == "6.00"
    assert len(invoice["lines"]) == 2

    exported = client.get(f"/api/v1/billing/me/invoices/download?invoiceIds={invoice['id']}")
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert exported.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    sheet = workbook.active
    assert sheet["F2"].value == "INV26001"
    assert sheet["G2"].value == "2026.08.17"
    assert "Customer Limited" in sheet["A7"].value
    assert {sheet["A13"].value, sheet["A14"].value} == {first.waybill_number, second.waybill_number}
    assert sheet["A15"].value is None
    assert sheet["D15"].value is None
    assert sheet["G17"].value is None
    assert sheet["G15"].value == 6
    assert "Beneficiary Name: Epix Logistics Co., Limited" in sheet["A17"].value
    assert sheet["A20"].value is None
    # The reference template already contains two decorative images; exporting
    # the invoice adds the configured, semi-transparent seal as another image.
    assert len(sheet._images) >= 3
    bank_bottom_row = 19
    stamp_anchor = _compact_stamp_anchor({"anchor": "B34"}, 15, 17, 300, sheet)
    stamp_start_row = int(stamp_anchor[1:])
    stamp_available_height = sum(
        (sheet.row_dimensions[row].height or sheet.sheet_format.defaultRowHeight or 15) * 96 / 72
        for row in range(stamp_start_row, bank_bottom_row + 1)
    )
    assert stamp_start_row <= 16
    assert stamp_available_height >= 300
    workbook.close()

    assert login(client, email=admin.email).status_code == 200
    voided = client.post(f"/api/v1/billing/users/{user.id}/invoices/{invoice['id']}/void", json={"reason": "Correction"})
    assert voided.status_code == 200
    assert voided.json()["status"] == "voided"
    available_again = client.get(f"/api/v1/billing/users/{user.id}/invoices/eligible")
    assert {item["id"] for item in available_again.json()} == {str(first.id), str(second.id)}

    saved = db_session.execute(select(Invoice)).scalar_one()
    assert saved.payer_snapshot["companyName"] == "Customer Limited"


def test_more_than_thirty_deductions_create_two_invoices(client, db_session, tmp_path):
    admin = create_test_user(db_session, email="admin@example.com", username="Admin", role="admin")
    user = create_test_user(db_session, email="customer@example.com", username="Customer")
    user.payer_company_name = "Customer Limited"
    user.payer_address_info = "Customer address"
    ready_settings(db_session, tmp_path, admin)
    entries = [deduction(db_session, user, admin, index) for index in range(31)]
    db_session.commit()
    assert login(client, email=admin.email).status_code == 200

    created = client.post(
        f"/api/v1/billing/users/{user.id}/invoices",
        json={"deductionIds": [str(entry.id) for entry in entries], "issuedDate": date(2026, 8, 17).isoformat()},
    )
    assert created.status_code == 201
    invoices = created.json()["invoices"]
    assert [item["invoiceNumber"] for item in invoices] == ["INV26001", "INV26002"]
    assert [len(item["lines"]) for item in invoices] == [30, 1]

    batch = client.get(
        f"/api/v1/billing/users/{user.id}/invoices/download?invoiceIds={invoices[0]['id']}&invoiceIds={invoices[1]['id']}"
    )
    assert batch.status_code == 200
    assert batch.headers["content-type"].startswith("application/zip")
    with ZipFile(BytesIO(batch.content)) as archive:
        compact_workbook = load_workbook(BytesIO(archive.read("INV26002.xlsx")), data_only=True)
    compact_sheet = compact_workbook.active
    assert compact_sheet["A13"].value == invoices[1]["lines"][0]["waybillNumber"]
    assert compact_sheet["G14"].value == 3
    assert "Beneficiary Name: Epix Logistics Co., Limited" in compact_sheet["A16"].value
    compact_workbook.close()


def test_admin_updates_each_invoice_setting_independently(client, db_session):
    admin = create_test_user(db_session, email="admin@example.com", username="Admin", role="admin")
    assert login(client, email=admin.email).status_code == 200

    first = client.patch("/api/v1/billing/invoice-settings", json={"beneficiaryName": "EPIX"})
    assert first.status_code == 200
    assert first.json()["beneficiaryName"] == "EPIX"
    assert first.json()["bankAccount"] is None

    second = client.patch("/api/v1/billing/invoice-settings", json={"bankAccount": "7949929686"})
    assert second.status_code == 200
    assert second.json()["beneficiaryName"] == "EPIX"
    assert second.json()["bankAccount"] == "7949929686"


def test_admin_can_preview_configured_invoice_stamp(client, db_session, tmp_path):
    admin = create_test_user(db_session, email="admin@example.com", username="Admin", role="admin")
    ready_settings(db_session, tmp_path, admin)
    db_session.commit()
    assert login(client, email=admin.email).status_code == 200

    response = client.get("/api/v1/billing/invoice-settings/stamp")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("image/png")
    assert response.content.startswith(b"\x89PNG")
