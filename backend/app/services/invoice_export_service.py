from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.cell import coordinate_from_string

from app.db.models import Invoice


TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "assets" / "invoice_template.xlsx"
DETAIL_START_ROW = 13
DETAIL_MAX_LINES = 30


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, 8):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def _prepare_detail_area(sheet, line_count: int) -> tuple[int, int]:
    """Resize the template detail table to exactly the invoice's line count."""
    if not 1 <= line_count <= DETAIL_MAX_LINES:
        raise ValueError(f"Invoice detail count must be between 1 and {DETAIL_MAX_LINES}")

    template_detail_lines = 26
    # The bank block may be shifted when rows are inserted or removed, so release
    # the merge before changing the detail table's height.
    for merged in list(sheet.merged_cells.ranges):
        if str(merged) == "A41:G43":
            sheet.unmerge_cells(str(merged))

    # The supplied template contains example waybills (including "duty" rows).
    # Preserve its table formatting while removing every example value/formula.
    for row in range(DETAIL_START_ROW, DETAIL_START_ROW + template_detail_lines):
        for column in range(1, 8):
            sheet.cell(row, column).value = None

    if line_count > template_detail_lines:
        extra_lines = line_count - template_detail_lines
        # The template footer begins at row 39.  Insert enough formatted detail
        # rows directly before it for the 27th through 30th waybill.
        sheet.insert_rows(39, amount=extra_lines)
        for row in range(39, 39 + extra_lines):
            _copy_row_style(sheet, 38, row)
    elif line_count < template_detail_lines:
        # Remove unused rows rather than leaving a fixed 30-row blank area.
        sheet.delete_rows(DETAIL_START_ROW + line_count, template_detail_lines - line_count)

    total_row = DETAIL_START_ROW + line_count
    bank_row = total_row + 2
    # Recreate the three-row bank block immediately below the compact footer.
    for row in range(bank_row, bank_row + 3):
        for column in range(1, 8):
            sheet.cell(row, column).value = None
    sheet.merge_cells(start_row=bank_row, start_column=1, end_row=bank_row + 2, end_column=7)
    return total_row, bank_row


def _row_height_pixels(sheet, row: int) -> float:
    """Return an Excel row's displayed height in pixels."""
    points = sheet.row_dimensions[row].height or sheet.sheet_format.defaultRowHeight or 15
    return points * 96 / 72


def _image_anchor_row(image) -> int:
    """Get an embedded image's 1-based start row from either anchor type."""
    anchor = image.anchor
    if isinstance(anchor, str):
        _, row = coordinate_from_string(anchor)
        return row
    marker = getattr(anchor, "_from", None)
    return marker.row + 1 if marker is not None else 1


def _remove_template_detail_images(sheet) -> None:
    """Discard the template's embedded detail-area seal before adding our own."""
    # The supplied template has a large pre-existing image anchored in its detail
    # section.  It is not the configured invoice stamp and can extend below bank
    # details, so retain only header images from the template.
    sheet._images = [image for image in sheet._images if _image_anchor_row(image) < DETAIL_START_ROW]


def _compact_stamp_anchor(position: dict, total_row: int, bank_row: int, image_height: int, sheet) -> str:
    """Keep a saved stamp near a compact invoice without changing its size."""
    anchor = str(position.get("anchor", "E20"))
    column, row = coordinate_from_string(anchor)
    # Leave the one template spacer row after the total.  When the saved anchor
    # was in an omitted detail row, the full-size stamp may overlap the bank
    # block; this is intentional and avoids shrinking the seal.
    desired_row = min(row, total_row + 1)

    # A seal must never extend below the three-row bank-details cell.  Find the
    # lowest starting row that can fit the saved image height above its bottom;
    # then move it upward only when that boundary requires it.
    bank_bottom_row = bank_row + 2
    available_height = 0.0
    maximum_start_row = bank_bottom_row
    while maximum_start_row > 1 and available_height < image_height:
        available_height += _row_height_pixels(sheet, maximum_start_row)
        maximum_start_row -= 1
    if available_height < image_height:
        maximum_start_row = 1
    else:
        maximum_start_row += 1
    return f"{column}{min(desired_row, maximum_start_row)}"


def _bank_text(snapshot: dict) -> str:
    return "\n".join(
        [
            f"Beneficiary Name: {snapshot['beneficiaryName']}",
            f"Bank Account :{snapshot['bankAccount']}",
            f"Bank name and code: {snapshot['bankNameAndCode']}",
            f"Branch code: {snapshot['branchCode']}",
            f"Swift/BIC:{snapshot['swiftBic']}",
            f"Bank Address: {snapshot['bankAddress']}",
        ]
    )


def build_invoice_workbook(invoice: Invoice) -> bytes:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook.active
    total_row, bank_row = _prepare_detail_area(sheet, len(invoice.lines))
    _remove_template_detail_images(sheet)
    payer = invoice.payer_snapshot
    issuer = invoice.issuer_snapshot

    sheet["F2"] = invoice.invoice_number
    sheet["G2"] = invoice.issued_date.strftime("%Y.%m.%d")
    sheet["A7"] = f"{payer['companyName']}\n{payer['addressInfo']}"
    sheet["F7"] = f"{issuer['issuerCompanyName']}\n{issuer['issuerAddressInfo']}"
    sheet["A9"] = (
        f"賬單金額€ {invoice.total_amount:,.2f}  EUR"
        f"(到期日 {invoice.due_date.year}年{invoice.due_date.month}月{invoice.due_date.day}日)"
    )

    for index, line in enumerate(invoice.lines, start=DETAIL_START_ROW):
        sheet.cell(index, 1).value = line.waybill_number
        sheet.cell(index, 3).value = line.quantity
        sheet.cell(index, 4).value = float(line.unit_rate)
        sheet.cell(index, 7).value = float(line.amount)
        sheet.cell(index, 4).number_format = "0.00"
        sheet.cell(index, 7).number_format = "#,##0.00"

    sheet.cell(total_row, 6).value = "應付金額"
    sheet.cell(total_row, 7).value = float(invoice.total_amount)
    sheet.cell(total_row, 7).number_format = "#,##0.00"
    sheet.cell(bank_row, 1).value = _bank_text(issuer)

    if invoice.stamp_storage_path and invoice.stamp_position:
        stamp_path = Path(invoice.stamp_storage_path)
        if stamp_path.is_file():
            image = ExcelImage(stamp_path)
            image.width = int(invoice.stamp_position.get("width", 86))
            image.height = int(invoice.stamp_position.get("height", 86))
            sheet.add_image(image, _compact_stamp_anchor(invoice.stamp_position, total_row, bank_row, image.height, sheet))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
