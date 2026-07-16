import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Iterable


PARCEL_UNIT_NUMBER_COLUMN_INDEX = 9  # I
DESTINATION_COLUMN_INDEX = 19  # S
NUMBER_OF_ITEMS_COLUMN_INDEX = 21  # U
WEIGHT_KG_COLUMN_INDEX = 22  # V
MAX_NUMBER_OF_ITEMS = 20

COUNTRY_LOOKUP = {
    "AT": ("AT", "Austria"),
    "AUSTRIA": ("AT", "Austria"),
    "奥地利": ("AT", "Austria"),
    "BE": ("BE", "Belgium"),
    "BELGIUM": ("BE", "Belgium"),
    "比利时": ("BE", "Belgium"),
    "BG": ("BG", "Bulgaria"),
    "BULGARIA": ("BG", "Bulgaria"),
    "保加利亚": ("BG", "Bulgaria"),
    "CH": ("CH", "Switzerland"),
    "SWITZERLAND": ("CH", "Switzerland"),
    "瑞士": ("CH", "Switzerland"),
    "CZ": ("CZ", "Czechia"),
    "CZECHIA": ("CZ", "Czechia"),
    "CZECH REPUBLIC": ("CZ", "Czechia"),
    "捷克": ("CZ", "Czechia"),
    "DE": ("DE", "Germany"),
    "GERMANY": ("DE", "Germany"),
    "德国": ("DE", "Germany"),
    "DK": ("DK", "Denmark"),
    "DENMARK": ("DK", "Denmark"),
    "丹麦": ("DK", "Denmark"),
    "EE": ("EE", "Estonia"),
    "ESTONIA": ("EE", "Estonia"),
    "爱沙尼亚": ("EE", "Estonia"),
    "ES": ("ES", "Spain"),
    "SPAIN": ("ES", "Spain"),
    "西班牙": ("ES", "Spain"),
    "FI": ("FI", "Finland"),
    "FINLAND": ("FI", "Finland"),
    "芬兰": ("FI", "Finland"),
    "FR": ("FR", "France"),
    "FRANCE": ("FR", "France"),
    "法国": ("FR", "France"),
    "GB": ("GB", "United Kingdom"),
    "UK": ("GB", "United Kingdom"),
    "UNITED KINGDOM": ("GB", "United Kingdom"),
    "BRITAIN": ("GB", "United Kingdom"),
    "英国": ("GB", "United Kingdom"),
    "GR": ("GR", "Greece"),
    "GREECE": ("GR", "Greece"),
    "希腊": ("GR", "Greece"),
    "HR": ("HR", "Croatia"),
    "CROATIA": ("HR", "Croatia"),
    "CY": ("CY", "Cyprus"),
    "CYPRUS": ("CY", "Cyprus"),
    "克罗地亚": ("HR", "Croatia"),
    "HU": ("HU", "Hungary"),
    "HUNGARY": ("HU", "Hungary"),
    "匈牙利": ("HU", "Hungary"),
    "IE": ("IE", "Ireland"),
    "IRELAND": ("IE", "Ireland"),
    "爱尔兰": ("IE", "Ireland"),
    "IT": ("IT", "Italy"),
    "ITALY": ("IT", "Italy"),
    "意大利": ("IT", "Italy"),
    "LT": ("LT", "Lithuania"),
    "LITHUANIA": ("LT", "Lithuania"),
    "立陶宛": ("LT", "Lithuania"),
    "LU": ("LU", "Luxembourg"),
    "LUXEMBOURG": ("LU", "Luxembourg"),
    "卢森堡": ("LU", "Luxembourg"),
    "LV": ("LV", "Latvia"),
    "LATVIA": ("LV", "Latvia"),
    "MT": ("MT", "Malta"),
    "MALTA": ("MT", "Malta"),
    "拉脱维亚": ("LV", "Latvia"),
    "NL": ("NL", "Netherlands"),
    "NETHERLANDS": ("NL", "Netherlands"),
    "HOLLAND": ("NL", "Netherlands"),
    "荷兰": ("NL", "Netherlands"),
    "NO": ("NO", "Norway"),
    "NORWAY": ("NO", "Norway"),
    "挪威": ("NO", "Norway"),
    "PL": ("PL", "Poland"),
    "POLAND": ("PL", "Poland"),
    "波兰": ("PL", "Poland"),
    "PT": ("PT", "Portugal"),
    "PORTUGAL": ("PT", "Portugal"),
    "葡萄牙": ("PT", "Portugal"),
    "RO": ("RO", "Romania"),
    "ROMANIA": ("RO", "Romania"),
    "罗马尼亚": ("RO", "Romania"),
    "SE": ("SE", "Sweden"),
    "SWEDEN": ("SE", "Sweden"),
    "瑞典": ("SE", "Sweden"),
    "SI": ("SI", "Slovenia"),
    "SLOVENIA": ("SI", "Slovenia"),
    "斯洛文尼亚": ("SI", "Slovenia"),
    "SK": ("SK", "Slovakia"),
    "SLOVAKIA": ("SK", "Slovakia"),
    "斯洛伐克": ("SK", "Slovakia"),
}


class PreAlertParcelParseError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedPreAlertParcel:
    row_number: int
    parcel_unit_number: str
    number_of_items: int
    weight_kg: Decimal
    destination_raw: str
    destination_code: str | None
    destination_name: str | None


def parse_pre_alert_parcels(
    *,
    filename: str,
    content: bytes,
) -> list[ParsedPreAlertParcel]:
    rows = list(_read_rows(filename=filename, content=content))
    parcels = []
    errors = []
    for row_number, row in rows:
        parcel_unit_number = _cell_text(_cell(row, PARCEL_UNIT_NUMBER_COLUMN_INDEX))
        if not parcel_unit_number:
            continue

        number_of_items = _parse_items(
            _cell(row, NUMBER_OF_ITEMS_COLUMN_INDEX),
            row_number,
        )
        weight_kg = _parse_weight(_cell(row, WEIGHT_KG_COLUMN_INDEX), row_number)
        destination_raw = _cell_text(_cell(row, DESTINATION_COLUMN_INDEX))
        destination_code, destination_name = _resolve_country(destination_raw)
        if isinstance(number_of_items, str):
            errors.append(number_of_items)
            continue
        if isinstance(weight_kg, str):
            errors.append(weight_kg)
            continue

        parcels.append(
            ParsedPreAlertParcel(
                row_number=row_number,
                parcel_unit_number=parcel_unit_number,
                number_of_items=number_of_items,
                weight_kg=weight_kg,
                destination_raw=destination_raw,
                destination_code=destination_code,
                destination_name=destination_name,
            )
        )

    if errors:
        raise PreAlertParcelParseError(
            "Pre Alert parcel parse failed: " + "; ".join(errors[:20])
        )
    return parcels


def _read_rows(*, filename: str, content: bytes) -> Iterable[tuple[int, tuple[object, ...]]]:
    extension = Path(filename).suffix.lower()
    try:
        if extension == ".xlsx":
            return _read_xlsx_rows(content)
        if extension == ".xls":
            return _read_xls_rows(content)
    except Exception as exc:
        raise PreAlertParcelParseError(
            "Upload Pre Alert File could not be read as an Excel workbook"
        ) from exc
    raise PreAlertParcelParseError("Upload Pre Alert File must be an Excel workbook")


def _read_xlsx_rows(content: bytes) -> list[tuple[int, tuple[object, ...]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        return [
            (row_number, tuple(row))
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),
                start=2,
            )
        ]
    finally:
        workbook.close()


def _read_xls_rows(content: bytes) -> list[tuple[int, tuple[object, ...]]]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [
        (row_index + 1, tuple(sheet.row_values(row_index)))
        for row_index in range(1, sheet.nrows)
    ]


def _cell(row: tuple | list, one_based_index: int):
    index = one_based_index - 1
    return row[index] if len(row) > index else None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_items(value: object, row_number: int) -> int | str:
    decimal_value = _parse_decimal(value)
    if decimal_value is None or decimal_value != decimal_value.to_integral_value():
        return f"U row {row_number} must be an integer between 0 and {MAX_NUMBER_OF_ITEMS}"
    parsed = int(decimal_value)
    if parsed < 0 or parsed > MAX_NUMBER_OF_ITEMS:
        return f"U row {row_number} value must be between 0 and {MAX_NUMBER_OF_ITEMS}"
    return parsed


def _parse_weight(value: object, row_number: int) -> Decimal | str:
    decimal_value = _parse_decimal(value)
    if decimal_value is None or decimal_value < 0:
        return f"V row {row_number} weight must be a non-negative number"
    return decimal_value


def _parse_decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"(?i)\bkg\b|\beur\b|€", "", text).strip()
    text = text.replace(" ", "")
    if "," in text and "." not in text and text.count(",") == 1:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _resolve_country(value: str) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    key = re.sub(r"\s+", " ", value.strip()).upper()
    return COUNTRY_LOOKUP.get(key, (None, None))
