import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Iterable


RECIPIENT_NAME_COLUMN_INDEX = 12  # L
ADDRESS_COLUMN_INDEX = 13  # M
LIMITED_QUANTITY_COLUMN_INDEX = 21  # U
VALUE_EUR_COLUMN_INDEX = 23  # W
CONSISTENT_COLUMN_INDEXES = (1, 2, 3, 4, 5, 6, 7)  # A-G
MUST_BE_EMPTY_COLUMN_INDEXES = (14, 15, 17, 29)  # N, O, Q, AC
MAX_RECIPIENT_ADDRESS_VALUE_EUR = Decimal("150")
MAX_LIMITED_QUANTITY = Decimal("20")


class PreAlertValidationError(ValueError):
    pass


@dataclass(frozen=True)
class PreAlertRow:
    row_number: int
    values: tuple[object, ...]
    recipient_name: str
    address: str
    limited_quantity_raw: object
    value_eur_raw: object


def validate_pre_alert_excel(
    *,
    filename: str,
    content: bytes,
) -> None:
    rows = list(_read_rows(filename=filename, content=content))
    active_rows = [row for row in rows if not _is_blank_row(row)]
    validation_errors = [
        *_find_limited_quantity_errors(active_rows),
        *_find_must_be_empty_errors(active_rows),
        *_find_consistent_column_errors(active_rows),
        *_find_recipient_address_value_errors(active_rows),
    ]
    if validation_errors:
        raise PreAlertValidationError(
            "Pre Alert validation failed: " + "; ".join(validation_errors[:20])
        )


def _read_rows(*, filename: str, content: bytes) -> Iterable[PreAlertRow]:
    extension = Path(filename).suffix.lower()
    try:
        if extension == ".xlsx":
            return _read_xlsx_rows(content)
        if extension == ".xls":
            return _read_xls_rows(content)
    except Exception as exc:
        raise PreAlertValidationError(
            "Upload Pre Alert File could not be read as an Excel workbook"
        ) from exc

    raise PreAlertValidationError("Upload Pre Alert File must be an Excel workbook")


def _read_xlsx_rows(content: bytes) -> list[PreAlertRow]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        return [
            _build_row(row_number, row)
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),
                start=2,
            )
        ]
    finally:
        workbook.close()


def _read_xls_rows(content: bytes) -> list[PreAlertRow]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [
        _build_row(row_index + 1, sheet.row_values(row_index))
        for row_index in range(1, sheet.nrows)
    ]


def _build_row(row_number: int, row: tuple | list) -> PreAlertRow:
    return PreAlertRow(
        row_number=row_number,
        values=tuple(row),
        recipient_name=_cell_text(_cell(row, RECIPIENT_NAME_COLUMN_INDEX)),
        address=_cell_text(_cell(row, ADDRESS_COLUMN_INDEX)),
        limited_quantity_raw=_cell(row, LIMITED_QUANTITY_COLUMN_INDEX),
        value_eur_raw=_cell(row, VALUE_EUR_COLUMN_INDEX),
    )


def _cell(row: tuple | list, one_based_index: int):
    index = one_based_index - 1
    return row[index] if len(row) > index else None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_recipient_address_value_errors(rows: list[PreAlertRow]) -> list[str]:
    grouped: dict[tuple[str, str], list[tuple[PreAlertRow, Decimal]]] = defaultdict(list)
    errors = []

    for row in rows:
        raw_amount = _cell_text(row.value_eur_raw)
        if not row.recipient_name and not row.address and not raw_amount:
            continue

        amount = _parse_decimal(row.value_eur_raw)
        if amount is None:
            if raw_amount:
                errors.append(f"L/M/W row {row.row_number} amount is not a valid number")
            continue

        if not row.recipient_name or not row.address:
            errors.append(
                f"L/M/W row {row.row_number} recipient name and address are required when amount is present"
            )
            continue

        grouped[(_normalize_group_key(row.recipient_name), _normalize_group_key(row.address))].append(
            (row, amount)
        )

    for group_rows in grouped.values():
        total = sum((amount for _, amount in group_rows), Decimal("0"))
        if total > MAX_RECIPIENT_ADDRESS_VALUE_EUR:
            first_row = group_rows[0][0]
            row_numbers = ", ".join(str(row.row_number) for row, _ in group_rows[:8])
            errors.append(
                "同一收件人/地址的 W 列申报金额超过 150 EUR: "
                f"rows {row_numbers}, recipient {first_row.recipient_name}, "
                f"address {first_row.address}, total {_format_decimal(total)} EUR"
            )

    return errors[:10]


def _find_limited_quantity_errors(rows: list[PreAlertRow]) -> list[str]:
    errors = []
    for row in rows:
        raw_quantity = _cell_text(row.limited_quantity_raw)
        if not raw_quantity:
            continue

        quantity = _parse_decimal(row.limited_quantity_raw)
        if quantity is None:
            errors.append(f"U row {row.row_number} value is not a valid number")
            continue
        if quantity > MAX_LIMITED_QUANTITY:
            errors.append(
                f"U row {row.row_number} value must be less than or equal to 20"
            )
    return errors[:10]


def _find_must_be_empty_errors(rows: list[PreAlertRow]) -> list[str]:
    errors = []
    for row in rows:
        filled_columns = [
            _column_letter(column_index)
            for column_index in MUST_BE_EMPTY_COLUMN_INDEXES
            if _cell_text(_cell(row.values, column_index))
        ]
        if filled_columns:
            errors.append(
                f"{'/'.join(filled_columns)} row {row.row_number} must be empty"
            )
    return errors[:10]


def _find_consistent_column_errors(rows: list[PreAlertRow]) -> list[str]:
    if not rows:
        return []

    reference = _consistent_values(rows[0])
    errors = []
    for row in rows[1:]:
        current = _consistent_values(row)
        if current != reference:
            errors.append(
                f"A/B/C/D/E/F/G row {row.row_number} values must match row {rows[0].row_number}"
            )
    return errors[:10]


def _parse_decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"(?i)\beur\b|€", "", text).strip()
    text = text.replace(" ", "")
    if "," in text and "." not in text and text.count(",") == 1:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _normalize_group_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _format_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _is_blank_row(row: PreAlertRow) -> bool:
    return all(not _cell_text(value) for value in row.values)


def _consistent_values(row: PreAlertRow) -> tuple[str, ...]:
    return tuple(
        _normalize_consistency_value(_cell(row.values, column_index))
        for column_index in CONSISTENT_COLUMN_INDEXES
    )


def _normalize_consistency_value(value: object) -> str:
    return re.sub(r"\s+", " ", _cell_text(value))


def _column_letter(one_based_index: int) -> str:
    letters = ""
    index = one_based_index
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
