import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from app.schemas.supplier import SupplierFieldRule, SupplierVersionConfig
from app.services.pre_alert_parcel_parser import COUNTRY_LOOKUP


MAX_SAVED_ISSUES = 200


class SupplierStructureError(ValueError):
    pass


@dataclass(frozen=True)
class SupplierValidationIssue:
    rule_key: str
    rule_name: str
    row_number: int
    column: str
    message: str
    raw_value: str

    def as_dict(self) -> dict:
        return {
            "ruleKey": self.rule_key,
            "ruleName": self.rule_name,
            "rowNumber": self.row_number,
            "column": self.column,
            "message": self.message,
            "rawValue": self.raw_value,
        }


@dataclass(frozen=True)
class ParsedSupplierParcel:
    parcel_unit_number: str
    number_of_items: int | None
    weight_kg: Decimal | None
    destination_raw: str | None
    destination_code: str | None
    destination_name: str | None


@dataclass(frozen=True)
class SupplierEvaluationResult:
    distinct_count: int
    issue_count: int
    issues: list[SupplierValidationIssue]
    parcels: list[ParsedSupplierParcel]


class SupplierRuleEngine:
    def evaluate(
        self,
        *,
        filename: str,
        content: bytes,
        config: SupplierVersionConfig,
    ) -> SupplierEvaluationResult:
        header, rows = self._read_workbook(filename=filename, content=content, config=config)
        indexes = self._resolve_indexes(header=header, config=config)
        field_map = {field.key: field for field in config.fields}
        row_key = field_map[config.row_key_field_key]
        group_field = field_map[config.billing_group_field_key]
        distinct_field = field_map[config.billing_distinct_field_key]

        active_rows = [
            (row_number, row)
            for row_number, row in rows
            if self._text(self._cell(row, indexes[row_key.key]))
        ]
        cartons_by_group: dict[str, set[str]] = {}
        for _, row in active_rows:
            group_value = self._normalize_distinct(
                self._cell(row, indexes[group_field.key]),
                case_insensitive=group_field.case_insensitive,
            )
            if not group_value:
                continue
            cartons = cartons_by_group.setdefault(group_value, set())
            carton_value = self._normalize_distinct(
                self._cell(row, indexes[distinct_field.key]),
                case_insensitive=distinct_field.case_insensitive,
            )
            if carton_value:
                cartons.add(carton_value)
        if not cartons_by_group:
            raise SupplierStructureError("Billing group field contains no valid values")
        billable_unit_count = sum(
            max(1, len(distinct_cartons))
            for distinct_cartons in cartons_by_group.values()
        )

        issues: list[SupplierValidationIssue] = []
        issue_count = 0
        seen_by_rule: dict[str, dict[str, int]] = {}
        parcels = []

        for row_number, row in active_rows:
            semantic_values: dict[str, Any] = {}
            for field in config.fields:
                raw_value = self._cell(row, indexes[field.key])
                parsed, messages = self._validate_value(field, raw_value)
                if field.constraints.unique and self._text(raw_value):
                    normalized = self._normalize_distinct(
                        raw_value,
                        case_insensitive=field.case_insensitive,
                    )
                    seen = seen_by_rule.setdefault(field.key, {})
                    if normalized in seen:
                        messages.append(
                            f"must be unique; first seen on row {seen[normalized]}"
                        )
                    else:
                        seen[normalized] = row_number
                for message in messages:
                    issue_count += 1
                    if len(issues) < MAX_SAVED_ISSUES:
                        issues.append(
                            SupplierValidationIssue(
                                rule_key=field.key,
                                rule_name=field.name,
                                row_number=row_number,
                                column=field.locator_value,
                                message=message,
                                raw_value=self._text(raw_value)[:160],
                            )
                        )
                if field.semantic_field:
                    semantic_values[field.semantic_field] = parsed

            parcel_number = semantic_values.get("parcel_unit_number")
            if parcel_number:
                destination = semantic_values.get("destination")
                parcels.append(
                    ParsedSupplierParcel(
                        parcel_unit_number=str(parcel_number),
                        number_of_items=semantic_values.get("number_of_items"),
                        weight_kg=semantic_values.get("weight_kg"),
                        destination_raw=(destination or {}).get("raw") if isinstance(destination, dict) else None,
                        destination_code=(destination or {}).get("code") if isinstance(destination, dict) else None,
                        destination_name=(destination or {}).get("name") if isinstance(destination, dict) else None,
                    )
                )

        return SupplierEvaluationResult(
            distinct_count=billable_unit_count,
            issue_count=issue_count,
            issues=issues,
            parcels=parcels,
        )

    def _read_workbook(
        self,
        *,
        filename: str,
        content: bytes,
        config: SupplierVersionConfig,
    ) -> tuple[tuple[object, ...], list[tuple[int, tuple[object, ...]]]]:
        extension = Path(filename).suffix.lower()
        try:
            if extension == ".xlsx":
                return self._read_xlsx(content, config)
            if extension == ".xls":
                return self._read_xls(content, config)
        except SupplierStructureError:
            raise
        except Exception as exc:
            raise SupplierStructureError(
                "Upload Pre Alert File could not be read as an Excel workbook"
            ) from exc
        raise SupplierStructureError("Upload Pre Alert File must be an XLS or XLSX workbook")

    def _read_xlsx(
        self, content: bytes, config: SupplierVersionConfig
    ) -> tuple[tuple[object, ...], list[tuple[int, tuple[object, ...]]]]:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        try:
            if config.workbook.sheet_mode == "named":
                name = config.workbook.sheet_name or ""
                if name not in workbook.sheetnames:
                    raise SupplierStructureError(f"Worksheet '{name}' was not found")
                sheet = workbook[name]
            else:
                sheet = workbook.active
            header = tuple(
                next(
                    sheet.iter_rows(
                        min_row=config.workbook.header_row,
                        max_row=config.workbook.header_row,
                        values_only=True,
                    ),
                    (),
                )
            )
            rows = [
                (row_number, tuple(row))
                for row_number, row in enumerate(
                    sheet.iter_rows(
                        min_row=config.workbook.data_start_row,
                        values_only=True,
                    ),
                    start=config.workbook.data_start_row,
                )
            ]
            return header, rows
        finally:
            workbook.close()

    def _read_xls(
        self, content: bytes, config: SupplierVersionConfig
    ) -> tuple[tuple[object, ...], list[tuple[int, tuple[object, ...]]]]:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=content)
        if config.workbook.sheet_mode == "named":
            name = config.workbook.sheet_name or ""
            if name not in workbook.sheet_names():
                raise SupplierStructureError(f"Worksheet '{name}' was not found")
            sheet = workbook.sheet_by_name(name)
        else:
            sheet = workbook.sheet_by_index(0)
        header_index = config.workbook.header_row - 1
        header = tuple(sheet.row_values(header_index)) if sheet.nrows > header_index else ()
        rows = [
            (index + 1, tuple(sheet.row_values(index)))
            for index in range(config.workbook.data_start_row - 1, sheet.nrows)
        ]
        return header, rows

    def _resolve_indexes(
        self,
        *,
        header: tuple[object, ...],
        config: SupplierVersionConfig,
    ) -> dict[str, int]:
        header_lookup: dict[str, list[int]] = {}
        for index, value in enumerate(header):
            key = self._text(value).casefold()
            if key:
                header_lookup.setdefault(key, []).append(index)

        indexes = {}
        for field in config.fields:
            if field.locator_mode == "column":
                indexes[field.key] = self._column_index(field.locator_value)
                continue
            key = field.locator_value.strip().casefold()
            matches = header_lookup.get(key, [])
            if not matches:
                raise SupplierStructureError(
                    f"Header '{field.locator_value}' for rule '{field.name}' was not found"
                )
            if len(matches) > 1:
                raise SupplierStructureError(
                    f"Header '{field.locator_value}' is duplicated"
                )
            indexes[field.key] = matches[0]
        return indexes

    def _validate_value(
        self, field: SupplierFieldRule, raw_value: object
    ) -> tuple[Any, list[str]]:
        text = self._text(raw_value)
        if not text:
            if field.blank_policy == "required":
                return None, ["is required"]
            return None, []

        messages = []
        parsed: Any = text
        numeric: Decimal | None = None
        if field.value_type in {"number", "integer"}:
            numeric = self._decimal(raw_value)
            if numeric is None:
                return None, [
                    "must be an integer" if field.value_type == "integer" else "must be a number"
                ]
            if field.value_type == "integer" and numeric != numeric.to_integral_value():
                return None, ["must be an integer"]
            parsed = int(numeric) if field.value_type == "integer" else numeric
        elif field.value_type == "country":
            parsed = self._country(field, text)
            if parsed["code"] is None and not field.allow_unknown_country:
                messages.append("must be a recognized country")

        constraints = field.constraints
        if numeric is not None:
            if constraints.min_value is not None and numeric < constraints.min_value:
                messages.append(f"must be greater than or equal to {constraints.min_value}")
            if constraints.max_value is not None and numeric > constraints.max_value:
                messages.append(f"must be less than or equal to {constraints.max_value}")
        if constraints.min_length is not None and len(text) < constraints.min_length:
            messages.append(f"must contain at least {constraints.min_length} characters")
        if constraints.max_length is not None and len(text) > constraints.max_length:
            messages.append(f"must contain at most {constraints.max_length} characters")
        if constraints.pattern and not re.fullmatch(constraints.pattern, text[:1000]):
            messages.append("does not match the required pattern")
        if constraints.allowed_values:
            normalized = text.casefold() if field.case_insensitive else text
            allowed = {
                value.casefold() if field.case_insensitive else value
                for value in constraints.allowed_values
            }
            if normalized not in allowed:
                messages.append("is not an allowed value")
        return parsed, messages

    def _country(self, field: SupplierFieldRule, value: str) -> dict:
        custom = {
            source.strip().casefold(): code
            for source, code in field.country_aliases.items()
        }
        if value.casefold() in custom:
            code = custom[value.casefold()]
            name = next(
                (name for candidate_code, name in COUNTRY_LOOKUP.values() if candidate_code == code),
                code,
            )
            return {"raw": value, "code": code, "name": name}
        key = re.sub(r"\s+", " ", value.strip()).upper()
        code, name = COUNTRY_LOOKUP.get(key, (None, None))
        return {"raw": value, "code": code, "name": name}

    def _normalize_distinct(self, value: object, *, case_insensitive: bool) -> str:
        text = self._text(value)
        return text.casefold() if case_insensitive else text

    def _decimal(self, value: object) -> Decimal | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float, Decimal)):
            try:
                return Decimal(str(value))
            except InvalidOperation:
                return None
        text = self._text(value)
        text = re.sub(r"(?i)\bkg\b|\beur\b|€", "", text).strip().replace(" ", "")
        if "," in text and "." not in text and text.count(",") == 1:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def _column_index(self, letters: str) -> int:
        index = 0
        for letter in letters.upper():
            index = index * 26 + ord(letter) - 64
        return index - 1

    def _cell(self, row: tuple[object, ...], index: int) -> object:
        return row[index] if len(row) > index else None

    def _text(self, value: object) -> str:
        return "" if value is None else str(value).strip()
