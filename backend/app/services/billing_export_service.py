from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.billing import BillingAccountResponse, BillingEntryItem


TITLE_FILL = PatternFill("solid", fgColor="0B3B38")
SECTION_FILL = PatternFill("solid", fgColor="DCEDEA")
HEADER_FILL = PatternFill("solid", fgColor="E9F3F1")
LABEL_FILL = PatternFill("solid", fgColor="F4F7F9")
REFUND_FILL = PatternFill("solid", fgColor="ECFDF5")
CANCELLED_FILL = PatternFill("solid", fgColor="FFF1F2")
WHITE_FONT = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Aptos", size=12, bold=True, color="0B3B38")
LABEL_FONT = Font(name="Aptos", size=9, bold=True, color="667085")
BODY_FONT = Font(name="Aptos", size=10, color="344054")
HEADER_FONT = Font(name="Aptos", size=9, bold=True, color="0B3B38")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D8E1E7"),
)
EUR_FORMAT = '€#,##0.00;[Red]-€#,##0.00'
DATE_FORMAT = "yyyy-mm-dd hh:mm"


def _safe_text(value: str | None) -> str:
    text = value or ""
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _entry_label(entry: BillingEntryItem) -> str:
    return (
        "Tax cancellation"
        if entry.entry_type == "deduction_reversal"
        else "Customs tax deduction"
    )


def _entry_status(entry: BillingEntryItem) -> str:
    if entry.entry_type == "deduction_reversal":
        return "Refunded"
    if entry.reversed_by_entry_id is not None:
        return "Cancelled"
    return "Posted"


def _source_label(entry: BillingEntryItem) -> str:
    if entry.entry_type == "deduction_reversal":
        return "Tax cancellation"
    if entry.billing_source == "retroactive":
        return "Tax backfill"
    return "Upload"


def _excel_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def build_billing_workbook(account: BillingAccountResponse) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customer Billing"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A9"

    sheet.merge_cells("A1:M1")
    sheet["A1"] = "EPIX · CUSTOMER BILLING"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = WHITE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells("A2:M2")
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    sheet["A2"] = f"Deduction ledger · Generated {generated_at}"
    sheet["A2"].font = Font(name="Aptos", size=10, color="667085")
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 22

    summary_blocks = (
        ("A4:C4", "A5:C5", "CUSTOMER", _safe_text(account.user.username)),
        ("D4:G4", "D5:G5", "EMAIL", _safe_text(account.user.email)),
        (
            "H4:J4",
            "H5:J5",
            "CURRENT BALANCE",
            Decimal(account.user.balance),
        ),
        ("K4:M4", "K5:M5", "RECORDS", len(account.deductions)),
    )
    for label_range, value_range, label, value in summary_blocks:
        sheet.merge_cells(label_range)
        sheet.merge_cells(value_range)
        label_cell = sheet[label_range.split(":")[0]]
        value_cell = sheet[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = LABEL_FILL
        label_cell.font = LABEL_FONT
        label_cell.alignment = Alignment(vertical="center")
        value_cell.value = value
        value_cell.font = Font(name="Aptos", size=12, bold=True, color="101828")
        value_cell.alignment = Alignment(vertical="center")
    sheet["H5"].number_format = EUR_FORMAT
    sheet["K5"].number_format = "#,##0"
    sheet.row_dimensions[4].height = 20
    sheet.row_dimensions[5].height = 27

    sheet.merge_cells("A7:M7")
    sheet["A7"] = "DEDUCTION ENTRIES"
    sheet["A7"].fill = SECTION_FILL
    sheet["A7"].font = SECTION_FONT
    sheet["A7"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[7].height = 25

    headers = [
        "Recorded At",
        "Air Waybill Number",
        "Entry",
        "Status",
        "Supplier",
        "Supplier Version",
        "Source",
        "Arrival Airport",
        "Billable Units",
        "Unit Rate (EUR)",
        "Amount (EUR)",
        "Balance After (EUR)",
        "Record ID",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=8, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.row_dimensions[8].height = 30

    for row_number, entry in enumerate(account.deductions, start=9):
        signed_amount = (
            Decimal(entry.amount)
            if entry.entry_type == "deduction_reversal"
            else -Decimal(entry.amount)
        )
        row = [
            _excel_datetime(entry.created_at),
            _safe_text(entry.waybill_number),
            _entry_label(entry),
            _entry_status(entry),
            _safe_text(entry.supplier_name) or "-",
            entry.supplier_version_number,
            _source_label(entry),
            _safe_text(entry.arrival_airport) or "-",
            entry.billable_unit_count,
            Decimal(entry.unit_rate) if entry.unit_rate is not None else None,
            signed_amount,
            Decimal(entry.balance_after),
            str(entry.id),
        ]
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")
            cell.border = THIN_BORDER
        sheet.cell(row=row_number, column=1).number_format = DATE_FORMAT
        for column in (10, 11, 12):
            sheet.cell(row=row_number, column=column).number_format = EUR_FORMAT
        if entry.entry_type == "deduction_reversal":
            for column in range(1, len(headers) + 1):
                sheet.cell(row=row_number, column=column).fill = REFUND_FILL
            sheet.cell(row=row_number, column=11).font = Font(
                name="Aptos", size=10, bold=True, color="047857"
            )
        elif entry.reversed_by_entry_id is not None:
            for column in range(1, len(headers) + 1):
                sheet.cell(row=row_number, column=column).fill = CANCELLED_FILL
        else:
            sheet.cell(row=row_number, column=11).font = Font(
                name="Aptos", size=10, bold=True, color="BE123C"
            )

    last_row = max(8, 8 + len(account.deductions))
    widths = [19, 22, 23, 13, 18, 16, 17, 16, 14, 16, 17, 20, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.auto_filter.ref = f"A8:M{last_row}"
    sheet.print_title_rows = "1:8"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
